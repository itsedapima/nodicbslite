import logging
from django.contrib import admin
from django.core.exceptions import ValidationError
from django.db import transaction, IntegrityError
from django.forms import ModelForm, CharField, ValidationError as FormValidationError
from django.utils.html import format_html
from django.urls import path
from django.template.response import TemplateResponse
from django.shortcuts import render, redirect
from django.http import HttpResponse
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

from .models import Customer, NextOfKin, CompanyBranch

logger = logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────────────────
#  FORMS WITH VALIDATION
# ──────────────────────────────────────────────────────────────────────

class CustomerAdminForm(ModelForm):
    """Custom form with integrity checks for Customer model."""
    
    class Meta:
        model = Customer
        fields = '__all__'
    
    def clean(self):
        """Validate Customer data before save."""
        cleaned_data = super().clean()
        
        # ── Required fields ────────────────────────────────────────
        if not cleaned_data.get('full_name'):
            raise FormValidationError("Full name is required.")
        
        if not cleaned_data.get('phone'):
            raise FormValidationError("Phone number is required.")
        
        if not cleaned_data.get('national_id'):
            raise FormValidationError("National ID is required.")
        
        # ── Uniqueness checks (excluding current instance on update) ─
        phone = cleaned_data.get('phone')
        national_id = cleaned_data.get('national_id')
        instance_pk = self.instance.pk
        
        # Check phone uniqueness
        if phone:
            phone_exists = Customer.objects.filter(phone=phone).exclude(pk=instance_pk).exists()
            if phone_exists:
                raise FormValidationError(f"Phone number '{phone}' is already registered.")
        
        # Check national_id uniqueness
        if national_id:
            nid_exists = Customer.objects.filter(national_id=national_id).exclude(pk=instance_pk).exists()
            if nid_exists:
                raise FormValidationError(f"National ID '{national_id}' is already registered.")
        
        # ── Logical validations ────────────────────────────────────
        dob = cleaned_data.get('dob')
        customer_type = cleaned_data.get('customer_type')
        
        if dob and customer_type == 'minor_individual':
            from datetime import date
            age = (date.today() - dob).days // 365
            if age >= 18:
                raise FormValidationError(
                    f"Selected customer type is 'Minor-Individual' but DOB indicates age {age}."
                )
        
        # ── Status consistency ─────────────────────────────────────
        customer_status = cleaned_data.get('customer_status')
        exit_date = cleaned_data.get('exit_date')
        death_date = cleaned_data.get('death_date')
        
        if customer_status == 'exited' and not exit_date:
            raise FormValidationError("Exit date is required when status is 'Exited'.")
        
        if customer_status == 'deceased' and not death_date:
            raise FormValidationError("Death date is required when status is 'Deceased'.")
        
        if exit_date and (exit_date > datetime.now().date()):
            raise FormValidationError("Exit date cannot be in the future.")
        
        if death_date and (death_date > datetime.now().date()):
            raise FormValidationError("Death date cannot be in the future.")
        
        return cleaned_data


class NextOfKinAdminForm(ModelForm):
    """Custom form with integrity checks for NextOfKin model."""
    
    class Meta:
        model = NextOfKin
        fields = '__all__'
    
    def clean(self):
        """Validate NextOfKin data before save."""
        cleaned_data = super().clean()
        
        # ── Required fields ────────────────────────────────────────
        if not cleaned_data.get('customer'):
            raise FormValidationError("Customer is required.")
        
        if not cleaned_data.get('kin_name'):
            raise FormValidationError("Kin name is required.")
        
        if not cleaned_data.get('kin_relationship'):
            raise FormValidationError("Kin relationship is required.")
        
        # ── Validate dates ─────────────────────────────────────────
        kin_dob = cleaned_data.get('kin_dob')
        if kin_dob and kin_dob > datetime.now().date():
            raise FormValidationError("Kin date of birth cannot be in the future.")
        
        # ── Phone validation (if provided) ─────────────────────────
        kin_phone = cleaned_data.get('kin_phone')
        if kin_phone and not kin_phone.isdigit():
            raise FormValidationError("Kin phone must contain only digits.")
        
        return cleaned_data


# ──────────────────────────────────────────────────────────────────────
#  INLINE ADMIN
# ──────────────────────────────────────────────────────────────────────

class NextOfKinInline(admin.TabularInline):
    """Inline admin for NextOfKin to manage directly with Customer."""
    model = NextOfKin
    form = NextOfKinAdminForm
    extra = 1
    fields = ('kin_name', 'kin_relationship', 'gender', 'kin_phone', 'kin_national_id')
    
    def get_extra(self, request, obj=None, **kwargs):
        """Don't show extra empty row if customer already has kins."""
        return 0 if obj else 1


# ──────────────────────────────────────────────────────────────────────
#  CUSTOMER ADMIN
# ──────────────────────────────────────────────────────────────────────

@admin.register(Customer)
class CustomerAdmin(admin.ModelAdmin):
    """Admin interface for Customer model with upload and validation."""
    
    form = CustomerAdminForm
    inlines = [NextOfKinInline]
    
    # ── Display configuration ──────────────────────────────────────
    list_display = (
        'cust_no', 'full_name', 'phone', 'customer_status',
        'customer_type', 'branch', 'reg_date', 'status_badge'
    )
    list_filter = (
        'customer_status', 'customer_type', 'branch', 'reg_date',
        'gender', 'marital_status', 'is_treasury'
    )
    search_fields = (
        'cust_no', 'full_name', 'phone', 'national_id', 'reg_email'
    )
    
    # ── Read-only fields ───────────────────────────────────────────
    readonly_fields = (
        'cust_no', 'reg_date', 'cust_no_int', 'display_name',
        'created_by', 'updated_by'
    )
    
    # ── Fieldset organization ──────────────────────────────────────
    fieldsets = (
        ('Member Number', {
            'fields': ('cust_no', 'cust_no_int'),
            'description': 'Auto-generated, zero-padded member number.',
        }),
        ('Authentication', {
            'fields': ('user',),
        }),
        ('Demographics', {
            'fields': (
                'full_name', 'first_name', 'middle_name', 'last_name',
                'customer_type', 'gender', 'marital_status', 'dob',
            ),
        }),
        ('Contact Information', {
            'fields': (
                'phone', 'reg_email', 'postal_address', 'postal_code',
                'home_address', 'town',
            ),
        }),
        ('Identity', {
            'fields': ('national_id', 'kra_pin'),
        }),
        ('Registration', {
            'fields': ('branch', 'reg_date', 'registered_by', 'reg_fee_is_paid'),
        }),
        ('Status', {
            'fields': (
                'customer_status', 'is_treasury', 'exit_date', 'exit_reason',
                'death_date', 'is_reactivated', 'reactivation_date',
                'reactivation_reason',
            ),
        }),
        ('Audit', {
            'fields': ('created_by', 'updated_by'),
            'classes': ('collapse',),
        }),
    )
    
    # ── Actions and custom URLs ────────────────────────────────────
    def get_urls(self):
        """Add custom admin URLs for bulk import."""
        urls = super().get_urls()
        custom_urls = [
            path('import-csv/', self.admin_site.admin_view(self.import_csv_view),
                 name='import_customers_csv'),
        ]
        return custom_urls + urls
    
    # ── Custom display methods ─────────────────────────────────────
    
    def status_badge(self, obj):
        """Display status with color-coded badge."""
        status_colors = {
            'active': 'green',
            'dormant': 'orange',
            'exited': 'red',
            'deceased': 'gray',
            'suspended': 'red',
        }
        color = status_colors.get(obj.customer_status, 'blue')
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; '
            'border-radius: 3px; font-weight: bold;">{}</span>',
            color,
            obj.get_customer_status_display()
        )
    status_badge.short_description = 'Status'
    
    def display_name(self, obj):
        """Show formatted display name."""
        return obj.display_name
    display_name.short_description = 'Display Name'
    
    # ── Bulk import functionality ──────────────────────────────────
    
    def import_csv_view(self, request):
        """Handle Excel import with validation and error reporting."""
        
        if request.method == 'POST' and request.FILES.get('excel_file'):
            return self._process_excel_import(request)
        
        # GET: Show import form
        return TemplateResponse(request, 'admin/customer_import.html', {
            'title': 'Import Customers from Excel',
            'site_header': self.admin_site.site_header,
            'has_add_permission': self.has_add_permission(request),
        })
    
    def _process_excel_import(self, request):
        """Parse Excel and bulk insert with integrity checks."""
        excel_file = request.FILES['excel_file']
        results = {
            'created': 0,
            'updated': 0,
            'errors': [],
            'warnings': [],
        }
        
        try:
            # ── Load and parse Excel ───────────────────────────────
            workbook = openpyxl.load_workbook(excel_file)
            worksheet = workbook.active
            
            if not worksheet or worksheet.max_row < 2:
                results['errors'].append("Excel file is empty or has no data rows.")
                return self._render_import_results(request, results)
            
            # ── Extract headers ────────────────────────────────────
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            if not headers:
                results['errors'].append("Excel file has no headers in first row.")
                return self._render_import_results(request, results)
            
            # ── Expected columns ───────────────────────────────────
            required_columns = {'full_name', 'phone', 'national_id'}
            header_set = set(headers)
            if not required_columns.issubset(header_set):
                missing = required_columns - header_set
                results['errors'].append(
                    f"Missing required columns: {', '.join(missing)}"
                )
                return self._render_import_results(request, results)
            
            # ── Process each row ───────────────────────────────────
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Build row dict from headers and values
                    row_dict = {}
                    for col_idx, header in enumerate(headers):
                        if col_idx < len(row):
                            value = row[col_idx]
                            row_dict[header] = value if value is not None else ''
                    
                    self._import_customer_row(row_dict, row_num, results, request.user)
                except Exception as e:
                    results['errors'].append(
                        f"Row {row_num}: {str(e)[:150]}"
                    )
            
            workbook.close()
            
        except Exception as e:
            results['errors'].append(f"File processing error: {str(e)[:150]}")
            logger.error(f"[CustomerAdmin] Excel import error: {e}")
        
        logger.info(
            f"[CustomerAdmin] Excel import completed: "
            f"{results['created']} created, {results['updated']} updated, "
            f"{len(results['errors'])} errors"
        )
        
        return self._render_import_results(request, results)
    
    def _import_customer_row(self, row, row_num, results, user):
        """Import a single customer row with validation."""
        
        # ── Extract and clean data ─────────────────────────────────
        full_name = row.get('full_name', '').strip()
        phone = row.get('phone', '').strip()
        national_id = row.get('national_id', '').strip()
        
        # Validate required fields
        if not full_name:
            raise ValueError("full_name is required and cannot be empty")
        if not phone:
            raise ValueError("phone is required and cannot be empty")
        if not national_id:
            raise ValueError("national_id is required and cannot be empty")
        
        # ── Optional fields ────────────────────────────────────────
        first_name = row.get('first_name', '').strip() or None
        middle_name = row.get('middle_name', '').strip() or None
        last_name = row.get('last_name', '').strip() or None
        gender = row.get('gender', '').strip() or None
        marital_status = row.get('marital_status', '').strip() or None
        postal_address = row.get('postal_address', '').strip() or None
        postal_code = row.get('postal_code', '').strip() or None
        home_address = row.get('home_address', '').strip() or None
        town = row.get('town', '').strip() or None
        customer_type = row.get('customer_type', 'adult_individual').strip()
        reg_email = row.get('reg_email', '').strip() or None
        kra_pin = row.get('kra_pin', '').strip() or None
        
        # Parse DOB
        dob = None
        dob_str = row.get('dob', '').strip()
        if dob_str:
            try:
                dob = datetime.strptime(dob_str, '%Y-%m-%d').date()
            except ValueError:
                raise ValueError(f"Invalid DOB format: '{dob_str}' (use YYYY-MM-DD)")
        
        # Parse branch
        branch = None
        branch_name = row.get('branch', '').strip()
        if branch_name:
            try:
                branch = CompanyBranch.objects.get(name=branch_name)
            except CompanyBranch.DoesNotExist:
                raise ValueError(
                    f"Branch '{branch_name}' not found. Available branches: "
                    f"{', '.join(CompanyBranch.objects.values_list('name', flat=True))}"
                )
        
        # ── Check for duplicates (before saving) ────────────────────
        phone_exists = Customer.objects.filter(phone=phone).exists()
        nid_exists = Customer.objects.filter(national_id=national_id).exists()
        
        if phone_exists and nid_exists:
            raise ValueError(
                f"Phone '{phone}' AND National ID '{national_id}' already exist in DB"
            )
        if phone_exists:
            raise ValueError(f"Phone '{phone}' already registered")
        if nid_exists:
            raise ValueError(f"National ID '{national_id}' already registered")
        
        # ── Save with transaction ──────────────────────────────────
        try:
            with transaction.atomic():
                customer = Customer(
                    full_name=full_name,
                    first_name=first_name,
                    middle_name=middle_name,
                    last_name=last_name,
                    phone=phone,
                    national_id=national_id,
                    gender=gender,
                    marital_status=marital_status,
                    postal_address=postal_address,
                    postal_code=postal_code,
                    home_address=home_address,
                    town=town,
                    customer_type=customer_type,
                    reg_email=reg_email,
                    kra_pin=kra_pin,
                    dob=dob,
                    branch=branch,
                    registered_by=user,
                    created_by=user.username,
                )
                customer.full_clean()  # Validate using model validators
                customer.save()
                results['created'] += 1
                
                logger.info(
                    f"[CustomerAdmin] Created customer {customer.cust_no} "
                    f"({full_name}) from CSV import"
                )
        
        except IntegrityError as e:
            raise ValueError(f"Database integrity error: {str(e)[:100]}")
    
    def _render_import_results(self, request, results):
        """Render results page after import."""
        return TemplateResponse(request, 'admin/customer_import_results.html', {
            'title': 'Import Results',
            'results': results,
            'total_processed': results['created'] + results['updated'] + len(results['errors']),
            'site_header': self.admin_site.site_header,
        })
    
    # ── Save override to track changes ─────────────────────────────
    
    def save_model(self, request, obj, form, change):
        """Override save to track user changes."""
        if not change:  # Creating new
            obj.created_by = request.user.username
        obj.updated_by = request.user.username
        super().save_model(request, obj, form, change)
    
    def save_formset(self, request, form, formset, change):
        """Override to track changes in related NextOfKin."""
        instances = formset.save(commit=False)
        for obj in instances:
            obj.updated_by = request.user.username
            if not obj.pk:
                obj.created_by = request.user.username
        super().save_formset(request, form, formset, change)


# ──────────────────────────────────────────────────────────────────────
#  NEXT OF KIN ADMIN
# ──────────────────────────────────────────────────────────────────────

@admin.register(NextOfKin)
class NextOfKinAdmin(admin.ModelAdmin):
    """Admin interface for NextOfKin model."""
    
    form = NextOfKinAdminForm
    
    list_display = (
        'kin_name', 'customer', 'kin_relationship', 'kin_phone'
    )
    list_filter = ('kin_relationship', 'gender', 'customer__branch')
    search_fields = ('kin_name', 'customer__full_name', 'kin_phone', 'kin_national_id')
    
    fieldsets = (
        ('Relationship', {
            'fields': ('customer', 'kin_relationship'),
        }),
        ('Personal Information', {
            'fields': ('kin_name', 'gender', 'kin_dob'),
        }),
        ('Contact & Identity', {
            'fields': ('kin_phone', 'kin_national_id'),
        }),
    )
    
    def get_queryset(self, request):
        """Optimize queryset with select_related."""
        return super().get_queryset(request).select_related('customer')