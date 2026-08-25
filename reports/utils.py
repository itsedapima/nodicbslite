"""
reports/utils.py
================
Shared helpers for all reports:

  - format_currency(...)           - pretty-print decimals
  - parse_date(...)                - safe date parser
  - get_date_range(request, ...)   - read ?start_date, ?end_date from GET
  - get_chama_info()               - chama logo/name/address for headers
  - render_excel(...)              - build a styled .xlsx HttpResponse
  - render_pdf(...)                - build a styled .pdf HttpResponse (reportlab)
  - render_report(...)             - dispatcher: html / excel / pdf based on
                                     ?format= query string.

Every report view should call render_report() once it has prepared
headers, rows, title, and (optionally) summary rows. The helper does
the rest.
"""

from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from typing import Iterable, List, Optional, Sequence, Tuple

from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from django.utils.dateparse import parse_date as _dj_parse_date


# -----------------------------------------------------------------------
#  Formatting helpers
# -----------------------------------------------------------------------

def format_currency(value, blank_zero: bool = False) -> str:
    """Return a comma-formatted 2-decimal string. '' on failure."""
    if value is None or value == "":
        return "" if blank_zero else "0.00"
    try:
        dec = Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return str(value)
    if blank_zero and dec == 0:
        return ""
    return f"{dec:,.2f}"


def parse_date(value, fallback=None):
    """Parse YYYY-MM-DD strings; return fallback on failure."""
    if value in (None, ""):
        return fallback
    if isinstance(value, (date, datetime)):
        return value.date() if isinstance(value, datetime) else value
    parsed = _dj_parse_date(str(value).strip())
    return parsed or fallback


def get_date_range(request, default_days: int = 30) -> Tuple[date, date]:
    """
    Read ?start_date= and ?end_date= from a GET request, returning a
    sane tuple of (start, end) dates.
    """
    today = timezone.localdate()
    start = parse_date(request.GET.get("start_date"), today - timedelta(days=default_days))
    end = parse_date(request.GET.get("end_date"), today)
    if start > end:
        start, end = end, start
    return start, end


def get_chama_info():
    """Lazily fetch the singleton ChamaInfo row for report headers."""
    try:
        from administration.models import ChamaInfo
        return ChamaInfo.objects.first()
    except Exception:
        return None


# -----------------------------------------------------------------------
#  EXCEL renderer
# -----------------------------------------------------------------------

def render_excel(
    filename: str,
    headers: Sequence[str],
    rows: Iterable[Sequence],
    *,
    title: str = "",
    subtitle: str = "",
    summary: Optional[Sequence[Tuple[str, str]]] = None,
    numeric_columns: Optional[Sequence[int]] = None,
) -> HttpResponse:
    """
    Build a styled .xlsx file and return it as an HttpResponse attachment.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Report")[:30]

    chama = get_chama_info()
    chama_name = (chama.chama_name if chama else "Chama MIS")

    # -- Header banner
    bold = Font(bold=True, size=14)
    sub_font = Font(italic=True, size=10, color="555555")
    thin = Side(border_style="thin", color="999999")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    fill_head = PatternFill("solid", fgColor="0E2B4D")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    fill_alt = PatternFill("solid", fgColor="F4F6F9")

    last_col = max(len(headers), 4)
    last_letter = get_column_letter(last_col)

    ws.merge_cells(f"A1:{last_letter}1")
    ws["A1"] = chama_name
    ws["A1"].font = bold
    ws["A1"].alignment = Alignment(horizontal="center")

    if title:
        ws.merge_cells(f"A2:{last_letter}2")
        ws["A2"] = title
        ws["A2"].font = Font(bold=True, size=12)
        ws["A2"].alignment = Alignment(horizontal="center")

    sub_line = subtitle or f"Generated: {timezone.localtime().strftime('%Y-%m-%d %H:%M')}"
    ws.merge_cells(f"A3:{last_letter}3")
    ws["A3"] = sub_line
    ws["A3"].font = sub_font
    ws["A3"].alignment = Alignment(horizontal="center")

    header_row_idx = 5

    # -- Column headers
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=str(h))
        cell.font = head_font
        cell.fill = fill_head
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    numeric_columns = set(numeric_columns or [])

    # -- Data rows
    row_idx = header_row_idx + 1
    for r_offset, row in enumerate(rows):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_coerce_cell(value))
            cell.border = border
            if r_offset % 2 == 1:
                cell.fill = fill_alt
            if (col_idx - 1) in numeric_columns:
                cell.alignment = Alignment(horizontal="right")
                if isinstance(cell.value, (int, float, Decimal)):
                    cell.number_format = "#,##0.00"
            else:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)
        row_idx += 1

    # -- Column widths
    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, min(40, len(str(h)) + 8))

    # -- Summary block
    if summary:
        row_idx += 1
        for label, value in summary:
            ws.cell(row=row_idx, column=1, value=str(label)).font = Font(bold=True)
            cell = ws.cell(row=row_idx, column=2, value=_coerce_cell(value))
            if isinstance(cell.value, (int, float, Decimal)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            cell.font = Font(bold=True)
            row_idx += 1

    # -- Stream out
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    safe = _slug(filename)
    response["Content-Disposition"] = f'attachment; filename="{safe}.xlsx"'
    return response


def _coerce_cell(value):
    """Make a value safe for openpyxl."""
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime,)):
        return value.replace(tzinfo=None)
    return value


def _slug(name: str) -> str:
    return "".join(c if c.isalnum() or c in "-_" else "_" for c in name).strip("_") or "report"


# -----------------------------------------------------------------------
#  PDF renderer  (reportlab)
# -----------------------------------------------------------------------

def render_pdf(
    filename: str,
    headers: Sequence[str],
    rows: Iterable[Sequence],
    *,
    title: str = "",
    subtitle: str = "",
    summary: Optional[Sequence[Tuple[str, str]]] = None,
    numeric_columns: Optional[Sequence[int]] = None,
    landscape_mode: bool = True,
) -> HttpResponse:
    """Build a styled A4 PDF (landscape by default)."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )

    chama = get_chama_info()
    chama_name = chama.chama_name if chama else "Chama MIS"
    chama_address = ""

    buf = io.BytesIO()
    page_size = landscape(A4) if landscape_mode else A4

    margin_h = 15 * mm
    margin_v = 15 * mm
    doc = SimpleDocTemplate(
        buf, pagesize=page_size,
        leftMargin=margin_h, rightMargin=margin_h,
        topMargin=margin_v, bottomMargin=margin_v,
        title=title or filename,
    )
    avail_width = page_size[0] - 2 * margin_h

    styles = getSampleStyleSheet()
    h_company = ParagraphStyle(
        "Company", parent=styles["Title"], fontSize=14, alignment=1,
        spaceAfter=2, textColor=colors.HexColor("#0E2B4D"),
    )
    h_addr = ParagraphStyle(
        "Addr", parent=styles["Normal"], fontSize=8, alignment=1,
        textColor=colors.grey, spaceAfter=4,
    )
    h_title = ParagraphStyle(
        "RptTitle", parent=styles["Heading2"], alignment=1, fontSize=12,
        spaceAfter=2,
    )
    h_sub = ParagraphStyle(
        "Sub", parent=styles["Normal"], alignment=1, fontSize=9,
        textColor=colors.grey, spaceAfter=8,
    )
    h_footer = ParagraphStyle(
        "Footer", parent=styles["Normal"], fontSize=7, alignment=1,
        textColor=colors.grey,
    )

    story = [
        Paragraph(chama_name, h_company),
        Paragraph(chama_address.replace("\n", " - "), h_addr),
        Paragraph(title or "Report", h_title),
        Paragraph(
            subtitle or f"Generated: {timezone.localtime().strftime('%Y-%m-%d %H:%M')}",
            h_sub,
        ),
    ]

    # -- Build table data
    numeric_columns = set(numeric_columns or [])
    data = [[str(h) for h in headers]]
    for r in rows:
        cells = []
        for col_idx, v in enumerate(r):
            if v is None:
                cells.append("")
                continue
            if (col_idx in numeric_columns) and isinstance(v, (int, float, Decimal)):
                cells.append(format_currency(v))
            elif isinstance(v, (date, datetime)):
                cells.append(v.strftime("%Y-%m-%d"))
            else:
                cells.append(str(v))
        data.append(cells)

    # -- Compute column widths to fill the page
    num_cols = len(headers)
    if num_cols > 0 and data:
        char_factor = 2.4 * mm
        min_num_w = 22 * mm
        min_txt_w = 18 * mm

        raw_widths = []
        for ci in range(num_cols):
            max_len = len(str(headers[ci]))
            for row in data[1:]:
                if ci < len(row):
                    max_len = max(max_len, len(str(row[ci])))
            desired = max(max_len * char_factor,
                          min_num_w if ci in numeric_columns else min_txt_w)
            raw_widths.append(desired)

        total_raw = sum(raw_widths) or 1
        col_widths = [max(w / total_raw * avail_width, 12 * mm) for w in raw_widths]

        scale = avail_width / (sum(col_widths) or 1)
        col_widths = [w * scale for w in col_widths]
    else:
        col_widths = None

    if num_cols <= 6:
        body_font_size = 9
    elif num_cols <= 10:
        body_font_size = 8.5
    else:
        body_font_size = 7.5

    table = Table(data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0E2B4D")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), body_font_size),
        ("FONTSIZE",   (0, 1), (-1, -1), body_font_size),
        ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",       (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F4F6F9")]),
        ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for col in numeric_columns:
        style_cmds.append(("ALIGN", (col, 1), (col, -1), "RIGHT"))
    table.setStyle(TableStyle(style_cmds))
    story.append(table)

    # -- Summary block
    if summary:
        story.append(Spacer(1, 10))
        sdata = [[str(lbl), str(val)] for lbl, val in summary]
        sum_w = [avail_width * 0.35, avail_width * 0.25]
        stable = Table(sdata, colWidths=sum_w, hAlign="RIGHT")
        stable.setStyle(TableStyle([
            ("FONTNAME",    (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 9),
            ("ALIGN",       (1, 0), (1, -1), "RIGHT"),
            ("BOX",         (0, 0), (-1, -1), 0.5, colors.HexColor("#0E2B4D")),
            ("INNERGRID",   (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",(0, 0), (-1, -1), 6),
            ("TOPPADDING",  (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
        ]))
        story.append(stable)

    story.append(Spacer(1, 10))
    story.append(Paragraph(
        f"{chama_name} - Report generated on "
        f"{timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')}",
        h_footer,
    ))

    doc.build(story)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type="application/pdf")
    safe = _slug(filename)
    response["Content-Disposition"] = f'attachment; filename="{safe}.pdf"'
    return response


# -----------------------------------------------------------------------
#  Unified dispatcher
# -----------------------------------------------------------------------

def render_report(
    request,
    *,
    template: str,
    title: str,
    headers: Sequence[str],
    rows: List[Sequence],
    subtitle: str = "",
    summary: Optional[Sequence[Tuple[str, str]]] = None,
    numeric_columns: Optional[Sequence[int]] = None,
    filename: Optional[str] = None,
    extra_context: Optional[dict] = None,
    landscape_mode: bool = True,
    page_size: int = 100,
):
    """
    Single entry-point used by every report view.

    GET parameters:
        format=excel  -> returns .xlsx (FULL dataset, no pagination)
        format=pdf    -> returns .pdf  (FULL dataset, no pagination)
        anything else -> renders the supplied HTML template

    HTML branch is paginated.
    """
    fmt = (request.GET.get("format") or "html").lower()
    fname = filename or _slug(title.lower())

    if fmt == "excel":
        return render_excel(
            fname, headers, rows,
            title=title, subtitle=subtitle, summary=summary,
            numeric_columns=numeric_columns,
        )
    if fmt == "pdf":
        return render_pdf(
            fname, headers, rows,
            title=title, subtitle=subtitle, summary=summary,
            numeric_columns=numeric_columns,
            landscape_mode=landscape_mode,
        )

    # -- HTML branch: paginate
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

    try:
        ps_override = int(request.GET.get("page_size", page_size))
        ps_override = max(10, min(ps_override, 1000))
    except (TypeError, ValueError):
        ps_override = page_size

    rows_list = list(rows)
    total_rows = len(rows_list)

    paginator = Paginator(rows_list, ps_override)
    page_no = request.GET.get("page") or 1
    try:
        page_obj = paginator.page(page_no)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages or 1)

    qd = request.GET.copy()
    qd.pop("page", None)
    base_query = qd.urlencode()

    context = {
        "report_title": title,
        "report_subtitle": subtitle,
        "headers": list(headers),
        "rows": list(page_obj.object_list),
        "all_rows_count": total_rows,
        "summary": list(summary) if summary else [],
        "numeric_columns": set(numeric_columns or []),
        "current_query": request.GET.urlencode(),
        "base_query": base_query,
        "page_obj": page_obj,
        "paginator": paginator,
        "is_paginated": paginator.num_pages > 1,
        "page_size": ps_override,
        "chama_info_obj": get_chama_info(),
    }
    if extra_context:
        context.update(extra_context)
    return render(request, template, context)


# -----------------------------------------------------------------------
#  Backwards-compatible HTML->Excel helper (kept from the original)
# -----------------------------------------------------------------------

def export_html_table_to_excel(html_table_content, file_name):
    """
    Parse an HTML table string and return an in-memory .xlsx BytesIO.
    """
    try:
        import pandas as pd
        dfs = pd.read_html(html_table_content)
        if not dfs:
            raise ValueError("No tables found in the provided HTML content.")
        df = dfs[0]
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
        buf.seek(0)
        return buf
    except Exception as e:
        print(f"Error during Excel export: {e}")
        return None
