from reportlab.platypus import Image, Spacer, Table
from reportlab.lib import colors
from administration.models import ChamaInfo  # Adjust import based on your app structure

def add_company_info(elements, width):
    """
    Adds company logo and details to the report.
    
    :param elements: List to which the elements will be appended.
    :param width: The total width of the report (to align text correctly).
    """
    # Fetch company Info
    company = ChamaInfo.objects.first()  # Assuming one company info record exists
    
    # Add company Logo (if available)
    if company and company.company_logo:
        logo_path = company.company_logo.path  # Get file path
        elements.append(Image(logo_path, width=80, height=80))
        elements.append(Spacer(1, 10))  # Add space after logo

    # company Details
    if company:
        company_details = f"{company.company_name}\n{company.company_address}\nTelephone: {company.company_contact}"
        elements.append(Table([[company_details]], colWidths=[width - 80], style=[
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 14),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse

def export_to_excel(filename, headers, data_rows):
    """
    A reusable utility to export data to an Excel file.
    :param filename: String name of the file (without extension)
    :param headers: List of strings for the top row
    :param data_rows: List of lists (the actual data)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Styling for Headers
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center")

    # Add Headers
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment

    # Add Data Rows
    for row in data_rows:
        ws.append(row)

    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Prepare Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response