"""
accounting/jv_service.py
=========================
Business logic for flexible Journal Vouchers with mixed customer /
SACCO member accounts on any leg.

A "line" is a dict with:
    entry_type       : 'sacco' | 'customer'
    cust_no          : str  (customer lines only)
    member_account   : str  (customer lines only — 'S01-00011', 'LN000037' ...)
    sacco_code       : str  (sacco lines only — '900-630010' ...)
    debit            : Decimal
    credit           : Decimal
    description      : str  (optional per-line note; else voucher description used)

validate_lines() returns:
    {
      'ok':   bool,
      'errors': [ {row: int, field: str, message: str}, ... ],
      'warnings': [ {row: int, message: str}, ... ],
      'resolved': [ {...enriched line...}, ... ],
      'total_debit':  Decimal,
      'total_credit': Decimal,
      'is_balanced':  bool,
    }

post_voucher() writes:
  1. JournalVoucher + JournalVoucherLine rows (status=posted)
  2. For sacco lines → GL only, via journal_entry()
  3. For customer savings-product lines → SavingsTransaction sub-ledger
     PLUS a GL leg on the product's linked sacco_gl_account.
  4. For customer loan lines → LoanTransaction sub-ledger PLUS a GL leg
     on the product's linked sacco_gl_account.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import List, Optional

from django.db import transaction as db_transaction
from django.utils import timezone

logger = logging.getLogger(__name__)

ZERO = Decimal("0.00")
TOLERANCE = Decimal("0.01")


def _dec(v) -> Decimal:
    if v in (None, "", "None"):
        return ZERO
    if isinstance(v, Decimal):
        return v
    try:
        return Decimal(str(v).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return ZERO


# ═════════════════════════════════════════════════════════════════════
#  LINE VALIDATION & RESOLUTION
# ═════════════════════════════════════════════════════════════════════

def validate_lines(lines: List[dict]) -> dict:
    """
    Validate a list of raw line dicts. Does NOT touch the database.
    Returns a report the UI can render errors from + the resolved lines
    (with FKs looked up) that post_voucher() can consume.
    """
    from accounting.models import SaccoAccount
    from customers.models import Customer
    from transactions.models import CustomerAccountsSetup

    errors: List[dict] = []
    warnings: List[dict] = []
    resolved: List[dict] = []
    total_dr = total_cr = ZERO

    if not lines:
        errors.append({"row": 0, "field": "lines",
                       "message": "At least two lines are required."})
        return {"ok": False, "errors": errors, "warnings": warnings,
                "resolved": [], "total_debit": ZERO, "total_credit": ZERO,
                "is_balanced": False}

    for i, raw in enumerate(lines, start=1):
        entry_type = (raw.get("entry_type") or "").strip().lower()
        dr = _dec(raw.get("debit"))
        cr = _dec(raw.get("credit"))
        description = (raw.get("description") or "").strip()

        # ── Amount checks ───────────────────────────────────────
        if dr < 0 or cr < 0:
            errors.append({"row": i, "field": "amount",
                           "message": "Amounts cannot be negative."})
            continue
        if dr > 0 and cr > 0:
            errors.append({"row": i, "field": "amount",
                           "message": "A line cannot have BOTH debit and credit."})
            continue
        if dr == 0 and cr == 0:
            errors.append({"row": i, "field": "amount",
                           "message": "A line needs a non-zero debit OR credit."})
            continue

        if entry_type not in ("sacco", "customer"):
            errors.append({"row": i, "field": "entry_type",
                           "message": "entry_type must be 'sacco' or 'customer'."})
            continue

        resolved_line = {
            "row": i,
            "entry_type": entry_type,
            "debit": dr,
            "credit": cr,
            "description": description,
        }

        # ── SACCO GL line ───────────────────────────────────────
        if entry_type == "sacco":
            code = (raw.get("sacco_code") or "").strip()
            if not code:
                errors.append({"row": i, "field": "sacco_code",
                               "message": "SACCO account code is required."})
                continue
            acct = SaccoAccount.objects.filter(account_code=code).first()
            if not acct:
                errors.append({"row": i, "field": "sacco_code",
                               "message": f"SACCO account '{code}' not found."})
                continue
            resolved_line["sacco_account"] = acct
            resolved_line["sacco_code"] = code
            resolved_line["display"] = f"{acct.account_code} — {acct.account_name}"

        # ── CUSTOMER member-account line ────────────────────────
        else:
            cust_no = (raw.get("cust_no") or "").strip()
            member_account = (raw.get("member_account") or "").strip()

            if not cust_no:
                errors.append({"row": i, "field": "cust_no",
                               "message": "Customer number is required."})
                continue
            if not member_account:
                errors.append({"row": i, "field": "member_account",
                               "message": "Member account is required for a customer line."})
                continue

            customer = Customer.objects.filter(cust_no=cust_no).first()
            if not customer:
                errors.append({"row": i, "field": "cust_no",
                               "message": f"Customer '{cust_no}' not found."})
                continue

            resolved_line["customer"] = customer
            resolved_line["cust_no"] = cust_no

            # Resolve member_account to either a savings product or a loan
            product, loan = _resolve_member_account(member_account, customer)

            if loan is not None:
                resolved_line["kind"] = "loan"
                resolved_line["loan"] = loan
                resolved_line["member_account_ref"] = loan.loan_no
                gl_code = loan.loan_type.get_gl_code() if loan.loan_type else None
                if not gl_code:
                    errors.append({"row": i, "field": "member_account",
                                   "message": (
                                       f"Loan product '{loan.loan_type.account_code}' "
                                       "has no linked GL account — configure "
                                       "'sacco_gl_account' in Admin."
                                   )})
                    continue
                gl_acct = SaccoAccount.objects.filter(account_code=gl_code).first()
                if not gl_acct:
                    errors.append({"row": i, "field": "member_account",
                                   "message": f"GL account '{gl_code}' not found."})
                    continue
                resolved_line["sacco_account"] = gl_acct
                resolved_line["display"] = (
                    f"{loan.loan_no} — Loan of {customer.full_name} "
                    f"(GL {gl_acct.account_code})"
                )

            elif product is not None:
                resolved_line["kind"] = "savings"
                resolved_line["product"] = product
                resolved_line["member_account_ref"] = (
                    f"{product.account_code}-{cust_no}"
                )
                gl_code = product.get_gl_code()
                if not gl_code:
                    errors.append({"row": i, "field": "member_account",
                                   "message": (
                                       f"Product '{product.account_code}' has no "
                                       "linked GL account — configure "
                                       "'sacco_gl_account' in Admin."
                                   )})
                    continue
                gl_acct = SaccoAccount.objects.filter(account_code=gl_code).first()
                if not gl_acct:
                    errors.append({"row": i, "field": "member_account",
                                   "message": f"GL account '{gl_code}' not found."})
                    continue
                resolved_line["sacco_account"] = gl_acct
                resolved_line["display"] = (
                    f"{product.account_code}-{cust_no} — "
                    f"{customer.full_name} · {product.account_name} "
                    f"(GL {gl_acct.account_code})"
                )

                # Balance warning for withdrawals
                if cr > 0 and getattr(product, 'is_withdrawable', True) and not product.is_loan_account:
                    balance = _get_savings_balance(customer.cust_no, product.account_type)
                    if cr > balance:
                        warnings.append({"row": i, "message": (
                            f"Withdrawal {cr:,.2f} exceeds current balance "
                            f"{balance:,.2f} on {product.account_name}. "
                            f"Posting will still proceed."
                        )})

            else:
                errors.append({"row": i, "field": "member_account",
                               "message": (
                                   f"Could not resolve member account "
                                   f"'{member_account}' for customer {cust_no}."
                               )})
                continue

        resolved.append(resolved_line)
        total_dr += dr
        total_cr += cr

    is_balanced = (total_dr - total_cr).copy_abs() <= TOLERANCE

    if resolved and not is_balanced:
        errors.append({"row": 0, "field": "balance", "message": (
            f"Journal is OUT OF BALANCE. Total debit KES {total_dr:,.2f} "
            f"≠ total credit KES {total_cr:,.2f} "
            f"(difference KES {(total_dr - total_cr).copy_abs():,.2f})."
        )})

    if len(resolved) < 2 and not errors:
        errors.append({"row": 0, "field": "lines",
                       "message": "At least two valid lines are required."})

    return {
        "ok": not errors,
        "errors": errors,
        "warnings": warnings,
        "resolved": resolved,
        "total_debit": total_dr,
        "total_credit": total_cr,
        "is_balanced": is_balanced,
    }


def _resolve_member_account(member_account_str: str, customer):
    """
    Given a member account string like 'S01-00011', 'S01', 'LN000037',
    'MOBI00000123', or 'normal_loan' — return either:
        (CustomerAccountsSetup, None)  for savings/deposit products
        (None, LoanHistory)            for a specific loan
        (None, None)                   if unresolvable
    """
    from transactions.models import CustomerAccountsSetup
    from loans.models import LoanHistory

    s = member_account_str.strip()

    # First, try as an explicit loan_no
    if s:
        loan = LoanHistory.objects.filter(
            loan_no=s, customer=customer,
        ).first()
        if loan:
            return None, loan

    # Try savings product by initials or account code
    # Format could be 'S01-00011' or 'S01' or just the account_type
    head = s.split("-")[0].strip()
    product = CustomerAccountsSetup.objects.filter(
        acc_initials=head, is_loan_account=False,
    ).first() or CustomerAccountsSetup.objects.filter(
        account_code=head, is_loan_account=False,
    ).first() or CustomerAccountsSetup.objects.filter(
        account_type=head, is_loan_account=False,
    ).first()
    if product:
        return product, None

    # Try loan product initials (e.g. 'MS' -> matches a mobile loan product)
    # then pick THE loan for this customer that matches the product
    loan_product = CustomerAccountsSetup.objects.filter(
        acc_initials=head, is_loan_account=True,
    ).first() or CustomerAccountsSetup.objects.filter(
        account_code=head, is_loan_account=True,
    ).first()
    if loan_product:
        loan = LoanHistory.objects.filter(
            customer=customer, loan_type=loan_product, is_disbursed=True,
        ).order_by("-loan_date").first()
        if loan:
            return None, loan

    return None, None


def _get_savings_balance(cust_no: str, account_type: str) -> Decimal:
    from django.db.models import Sum
    from transactions.models import SavingsTransaction
    agg = SavingsTransaction.objects.filter(
        cust_no=cust_no, saving_type=account_type,
    ).aggregate(cr=Sum("credit_amount"), dr=Sum("debit_amount"))
    return (agg["cr"] or ZERO) - (agg["dr"] or ZERO)


# ═════════════════════════════════════════════════════════════════════
#  POSTING
# ═════════════════════════════════════════════════════════════════════

def _next_voucher_no() -> str:
    from accounting.models import JournalVoucher
    last = JournalVoucher.objects.order_by("-id").first()
    n = 1
    if last and last.voucher_no and last.voucher_no.startswith("JV"):
        try:
            n = int(last.voucher_no[2:]) + 1
        except ValueError:
            n = 1
    return f"JV{n:07d}"


@dataclass
class PostResult:
    voucher_no: str
    voucher_id: int
    reference: str
    total_amount: Decimal
    line_count: int
    resolved_lines: list


def post_voucher(
    *,
    voucher_date,
    description: str,
    lines: List[dict],
    user,
    request=None,
) -> PostResult:
    """
    Validate and post a flexible journal voucher.
    Raises ValueError if validation fails.
    """
    from accounting.models import JournalVoucher, JournalVoucherLine
    from accounting.journal import journal_entry, leg
    from transactions.models import SavingsTransaction, LoanTransaction

    report = validate_lines(lines)
    if not report["ok"]:
        # Raise the first error and preserve the whole list for the view
        msg = report["errors"][0]["message"] if report["errors"] else "Validation failed."
        exc = ValueError(msg)
        exc.report = report
        raise exc

    resolved = report["resolved"]
    total_dr = report["total_debit"]

    with db_transaction.atomic():
        voucher_no = _next_voucher_no()
        voucher = JournalVoucher.objects.create(
            voucher_no=voucher_no,
            voucher_date=voucher_date or timezone.localdate(),
            description=description,
            status="posted",
            total_amount=total_dr,
            created_by=user if getattr(user, "pk", None) else None,
            approved_by=user if getattr(user, "pk", None) else None,
            approved_at=timezone.now(),
            posted_at=timezone.now(),
        )

        # Build GL legs and subledger writes
        gl_legs = []
        subledger_writes = []  # deferred so we know voucher_no

        for line in resolved:
            gl_acct = line["sacco_account"]
            dr = line["debit"]
            cr = line["credit"]

            # 1. Persist JournalVoucherLine row for audit/reprint
            jv_line = JournalVoucherLine(
                voucher=voucher,
                sacco_account=gl_acct,
                description=line["description"] or description,
                debit_amount=dr,
                credit_amount=cr,
                entry_type=line["entry_type"],
            )
            if line["entry_type"] == "customer":
                jv_line.customer = line["customer"]
                jv_line.member_account_ref = line.get("member_account_ref", "")
                if line.get("kind") == "savings":
                    jv_line.member_product = line["product"]
                elif line.get("kind") == "loan":
                    jv_line.member_loan_no = line["loan"].loan_no
            jv_line.save()

            # 2. GL leg (using accounting.journal format)
            gl_legs.append(leg(
                gl_acct,
                debit=dr, credit=cr,
                description=line["display"],
            ))

            # 3. Subledger writes for customer lines
            if line["entry_type"] == "customer":
                subledger_writes.append(line)

        # 4. Post the GL journal (this enforces balance + writes ledger)
        journal_entry(
            reference=voucher_no,
            description=description,
            legs=gl_legs,
            created_by=getattr(user, 'username', 'system') if user else 'system',
        )

        # 5. Now write the subledger rows
        #    tr_date MUST reflect the voucher_date (backdating support),
        #    NOT the moment of posting. voucher_date is a DateField; the
        #    subledger tr_date is a DateTimeField, so build an aware
        #    datetime at NOON of the voucher date (noon is timezone-shift
        #    safe — midnight can roll to the previous day under UTC storage).
        #    created_at (auto_now_add) still records the real posting moment.
        from datetime import datetime as _dt, time as _time
        _sub_tr_date = timezone.make_aware(
            _dt.combine(voucher.voucher_date, _time(12, 0))
        )

        for line in subledger_writes:
            customer = line["customer"]
            dr = line["debit"]
            cr = line["credit"]

            if line.get("kind") == "savings":
                product = line["product"]
                SavingsTransaction.objects.create(
                    cust_no=customer.cust_no,
                    saving_type=product.account_type,
                    tr_date=_sub_tr_date,
                    tr_ref=f"{voucher_no}-{line['row']}",
                    tr_desc=line["description"] or description,
                    debit_amount=dr,
                    credit_amount=cr,
                    created_by=getattr(user, "username", "system"),
                )
            elif line.get("kind") == "loan":
                loan = line["loan"]
                LoanTransaction.objects.create(
                    cust_no=customer.cust_no,
                    loan_id=loan.id,
                    loan_no=loan.loan_no,
                    loan_type=loan.loan_type.account_type,
                    tr_date=_sub_tr_date,
                    tr_ref=f"{voucher_no}-{line['row']}",
                    tr_desc=line["description"] or description,
                    debit_amount=dr,
                    credit_amount=cr,
                    created_by=getattr(user, "username", "system"),
                )

    return PostResult(
        voucher_no=voucher.voucher_no,
        voucher_id=voucher.id,
        reference=voucher.voucher_no,
        total_amount=total_dr,
        line_count=len(resolved),
        resolved_lines=resolved,
    )


# ═════════════════════════════════════════════════════════════════════
#  BULK IMPORT HELPERS
# ═════════════════════════════════════════════════════════════════════

TEMPLATE_HEADERS = [
    "entry_type", "cust_no", "member_account", "sacco_code",
    "debit", "credit", "description",
]


def parse_uploaded_lines(file_obj) -> List[dict]:
    """
    Parse a CSV or XLSX upload into a list of raw line dicts. Rejects
    empty rows silently. Extension detection is by filename.
    """
    import csv
    import io
    name = getattr(file_obj, "name", "").lower()

    lines: List[dict] = []

    if name.endswith(".xlsx") or name.endswith(".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ValueError("openpyxl is not installed; upload a CSV instead.")
        wb = load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return []
        headers = [str(h or "").strip().lower() for h in headers]
        for row in rows:
            if not row or all(c in (None, "") for c in row):
                continue
            rec = {h: (row[i] if i < len(row) else "") for i, h in enumerate(headers)}
            lines.append(_normalize_raw(rec))
    else:
        content = file_obj.read()
        if isinstance(content, bytes):
            content = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(content))
        for rec in reader:
            if not any((rec.get(k) or "").strip() for k in rec):
                continue
            lines.append(_normalize_raw(rec))

    return lines


def _normalize_raw(rec: dict) -> dict:
    """Coerce a parsed row to the raw-line shape validate_lines expects."""
    def _s(k):
        v = rec.get(k) or rec.get(k.upper()) or rec.get(k.capitalize()) or ""
        return str(v).strip() if v is not None else ""
    return {
        "entry_type": _s("entry_type").lower() or "sacco",
        "cust_no":    _s("cust_no"),
        "member_account": _s("member_account"),
        "sacco_code": _s("sacco_code"),
        "debit":  _s("debit")  or "0",
        "credit": _s("credit") or "0",
        "description": _s("description"),
    }


def build_template_csv() -> bytes:
    """Return a small CSV byte string demonstrating the upload format."""
    import csv, io
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(TEMPLATE_HEADERS)
    w.writerow(["sacco", "", "", "900-601001", "0", "1000.00",
                "Example: CR bank"])
    w.writerow(["customer", "00011", "S01", "", "1000.00", "0",
                "Example: DR member's Share Capital"])
    w.writerow(["customer", "00011", "LN000037", "", "0", "500.00",
                "Example: CR member's loan (repayment) — leave sacco_code blank"])
    w.writerow(["sacco", "", "", "900-600000", "500.00", "0",
                "Example: DR cash"])
    return buf.getvalue().encode("utf-8-sig")
