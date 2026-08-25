"""
data_imports/excel.py
=====================
Excel template generation and parsing helpers.

- build_template(import_type)  → BytesIO with a styled .xlsx template
- parse_file(file, columns)    → list[dict] with normalized keys

We deliberately use openpyxl (already in requirements) so we don't add
new dependencies.
"""
from __future__ import annotations

from io import BytesIO
from typing import Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
REQUIRED_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
EXAMPLE_FONT = Font(italic=True, color="7F7F7F")
THIN = Side(border_style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_template(import_type) -> BytesIO:
    """Build a styled .xlsx template for the given ImportType."""
    wb = Workbook()
    ws = wb.active
    ws.title = import_type.slug[:31]

    # ── Header row (row 1) ──────────────────────────────────────────
    for col_idx, col in enumerate(import_type.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col.header)
        cell.font = HEADER_FONT
        cell.fill = REQUIRED_FILL if col.required else HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        if col.help:
            marker = " (REQUIRED)" if col.required else ""
            cell.comment = Comment(f"{col.help}{marker}", "NODi Lite")

        # Set width heuristically
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            len(col.header) + 4, 18
        )

    ws.row_dimensions[1].height = 32

    # ── Example row (row 2, italic grey) ────────────────────────────
    for col_idx, col in enumerate(import_type.columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col.example)
        cell.font = EXAMPLE_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER

    # ── Freeze header, add filter ───────────────────────────────────
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = ws.dimensions

    # ── README sheet ────────────────────────────────────────────────
    readme = wb.create_sheet("README")
    readme.column_dimensions["A"].width = 24
    readme.column_dimensions["B"].width = 8
    readme.column_dimensions["C"].width = 80

    readme["A1"] = f"{import_type.title} — Import Template"
    readme["A1"].font = Font(bold=True, size=14, color="1F4E79")

    readme["A3"] = "Purpose"
    readme["A3"].font = Font(bold=True)
    readme["B3"] = ""
    readme["C3"] = import_type.description
    readme["C3"].alignment = Alignment(wrap_text=True, vertical="top")
    readme.row_dimensions[3].height = 60

    readme["A5"] = "Column"
    readme["B5"] = "Required"
    readme["C5"] = "Description"
    for c in ("A5", "B5", "C5"):
        readme[c].font = HEADER_FONT
        readme[c].fill = HEADER_FILL
        readme[c].alignment = Alignment(horizontal="center")

    r = 6
    for col in import_type.columns:
        readme.cell(row=r, column=1, value=col.header).font = Font(bold=True)
        readme.cell(row=r, column=2, value="YES" if col.required else "no").alignment = \
            Alignment(horizontal="center")
        readme.cell(row=r, column=3, value=col.help).alignment = \
            Alignment(wrap_text=True, vertical="top")
        readme.row_dimensions[r].height = 28
        r += 1

    readme.cell(row=r + 1, column=1, value="Notes").font = Font(bold=True)
    readme.cell(row=r + 2, column=3,
                value=("• Row 1 (blue) = column headers — do not rename or reorder.\n"
                       "• Row 2 (grey italic) = example row — DELETE it before uploading.\n"
                       "• Red headers = required.\n"
                       "• Dates: use YYYY-MM-DD (2024-06-15).\n"
                       "• Amounts: use plain numbers (no currency symbol, no commas)."))
    readme.cell(row=r + 2, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    readme.row_dimensions[r + 2].height = 100

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Parsing ────────────────────────────────────────────────────────────


def parse_file(uploaded_file, columns: List) -> List[Dict]:
    """Read an uploaded .xlsx (or .xls) into a list of dicts keyed by column header."""
    wb = load_workbook(uploaded_file, data_only=True, read_only=True)
    ws = wb.active

    headers_expected = [c.header for c in columns]
    headers_found: List[str] = []
    rows: List[Dict] = []

    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx == 1:
            # Normalize headers: strip, lowercase for matching
            headers_found = [str(h).strip() if h is not None else "" for h in row]
            continue

        # Skip completely empty rows
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        # Skip the example row: it repeats the exact "example" values, OR
        # any row whose cust_no equals a placeholder like "00116" AND matches
        # our own example row exactly. Simpler: don't skip; let validation
        # flag it if the SACCO doesn't have that member.
        # BUT we DO skip if the whole row equals the columns' example values.
        example_signature = tuple(str(c.example).strip() for c in columns)
        row_signature = tuple(str(v).strip() if v is not None else "" for v in row[:len(columns)])
        if example_signature == row_signature and any(example_signature):
            continue

        d = {}
        for i, header in enumerate(headers_found):
            if not header:
                continue
            key = header.strip()
            d[key] = row[i] if i < len(row) else None
        rows.append(d)

    # Validate that expected headers exist
    missing = [h for h in headers_expected if h not in headers_found]
    if missing:
        raise ValueError(
            f"Uploaded file is missing required column(s): {', '.join(missing)}. "
            f"Please download the template again."
        )

    return rows
